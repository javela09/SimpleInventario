import os
from datetime import datetime
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from flask import Flask, jsonify, render_template, request, send_file, session
from psycopg.rows import dict_row
from psycopg_pool import ConnectionPool


app = Flask(__name__)
_default_secret = "dev-secret-change-me"
app.secret_key = os.environ.get("SECRET_KEY", _default_secret)
if app.secret_key == _default_secret:
    app.logger.warning("SECRET_KEY no configurada; usando valor inseguro solo para desarrollo")

_pool: ConnectionPool | None = None
_schema_ready = False


def _normalize_database_url(raw_url: str) -> str:
    parsed = urlparse(raw_url)
    scheme = "postgresql" if parsed.scheme == "postgres" else parsed.scheme
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))

    # Neon exige conexiones TLS; si no se especifica, forzamos sslmode=require
    if "sslmode" not in query:
        query["sslmode"] = "require"

    normalized = parsed._replace(scheme=scheme, query=urlencode(query))
    return urlunparse(normalized)


def _ensure_schema(pool: ConnectionPool) -> None:
    """
    Crea tablas e indices necesarios de forma idempotente.
    Importante: usa pool.connection() directo para no llamar get_db()/get_pool() y evitar recursion.
    """
    global _schema_ready
    if _schema_ready:
        return

    with pool.connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS usuarios (
                    id SERIAL PRIMARY KEY,
                    nombre_usuario VARCHAR(255) NOT NULL UNIQUE,
                    es_admin BOOLEAN DEFAULT FALSE,
                    fecha_creacion TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
                """
            )

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS articulos (
                    id SERIAL PRIMARY KEY,
                    codigo_articulo VARCHAR(255) NOT NULL,
                    descripcion TEXT,
                    ean VARCHAR(255),
                    fecha_creacion TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
                """
            )

            # ON CONFLICT (ean) necesita un indice/constraint UNIQUE
            cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS articulos_ean_unique_idx ON articulos (ean)")

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS lecturas (
                    id SERIAL PRIMARY KEY,
                    usuario VARCHAR(255) NOT NULL,
                    ean VARCHAR(255) NOT NULL,
                    codigo_articulo VARCHAR(255) NOT NULL,
                    descripcion TEXT,
                    fecha_lectura TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
                """
            )

            # Usuarios admin por defecto (idempotente)
            cursor.execute(
                "INSERT INTO usuarios (nombre_usuario, es_admin) VALUES (%s, %s) ON CONFLICT (nombre_usuario) DO NOTHING",
                ("admin", True),
            )
            cursor.execute(
                "INSERT INTO usuarios (nombre_usuario, es_admin) VALUES (%s, %s) ON CONFLICT (nombre_usuario) DO NOTHING",
                ("henkobit", True),
            )

        conn.commit()

    _schema_ready = True


def get_pool() -> ConnectionPool:
    global _pool
    if _pool is None:
        database_url = os.environ.get("DATABASE_URL")
        if not database_url:
            raise RuntimeError("DATABASE_URL no configurada")

        conninfo = _normalize_database_url(database_url)
        _pool = ConnectionPool(
            conninfo=conninfo,
            min_size=int(os.environ.get("PGPOOL_MIN_SIZE", 1)),
            max_size=int(os.environ.get("PGPOOL_MAX_SIZE", 3)),
            kwargs={"row_factory": dict_row},
        )

    # Garantiza esquema (una sola vez por instancia)
    _ensure_schema(_pool)
    return _pool


def get_db():
    return get_pool().connection()


@app.route("/")
def index():
    return render_template("login.html")


@app.route("/api/login", methods=["POST"])
def login():
    data = request.json or {}
    usuario = (data.get("usuario") or "").strip()

    if not usuario:
        return jsonify({"success": False, "message": "Usuario requerido"}), 400

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, es_admin FROM usuarios WHERE nombre_usuario = %s", (usuario,))
            user = cursor.fetchone()

    if not user:
        return jsonify({"success": False, "message": "Usuario no autorizado"}), 403

    session["usuario"] = usuario
    session["es_admin"] = bool(user["es_admin"])
    return jsonify({"success": True, "es_admin": bool(user["es_admin"])})


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/lecturas")
def lecturas():
    if "usuario" not in session:
        return render_template("login.html")

    if session.get("es_admin"):
        return render_template("admin.html", usuario=session["usuario"])

    return render_template("lecturas.html", usuario=session["usuario"])


@app.route("/api/escanear", methods=["POST"])
def escanear():
    data = request.json or {}
    ean = (data.get("ean") or "").strip()

    if not ean:
        return jsonify({"success": False, "message": "Codigo de barras vacio"}), 400

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT codigo_articulo, descripcion FROM articulos WHERE ean = %s", (ean,))
            articulo = cursor.fetchone()

            if not articulo:
                return jsonify({"success": False, "message": f"El codigo {ean} no esta en el maestro"}), 404

            cursor.execute(
                """
                INSERT INTO lecturas (usuario, ean, codigo_articulo, descripcion)
                VALUES (%s, %s, %s, %s)
                RETURNING id
                """,
                (
                    session.get("usuario", "anonimo"),
                    ean,
                    articulo["codigo_articulo"],
                    articulo["descripcion"],
                ),
            )
            lectura_id = cursor.fetchone()["id"]
        conn.commit()

    return jsonify(
        {
            "success": True,
            "message": "Articulo encontrado y registrado",
            "lectura": {
                "id": lectura_id,
                "ean": ean,
                "codigo_articulo": articulo["codigo_articulo"],
                "descripcion": articulo["descripcion"],
            },
        }
    )


@app.route("/api/lecturas", methods=["GET"])
def obtener_lecturas():
    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, ean, codigo_articulo, descripcion, fecha_lectura
                FROM lecturas
                ORDER BY fecha_lectura DESC
                LIMIT 100
                """
            )
            lecturas = cursor.fetchall()
    return jsonify(lecturas)


@app.route("/api/lecturas/limpiar", methods=["DELETE"])
def limpiar_lecturas():
    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM lecturas")
        conn.commit()
    return jsonify({"success": True, "message": "Lecturas eliminadas"})


@app.route("/api/exportar", methods=["GET"])
def exportar_excel():
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT ean, codigo_articulo, descripcion, fecha_lectura
                FROM lecturas
                ORDER BY fecha_lectura DESC
                """
            )
            lecturas = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"
    headers = ["EAN", "Codigo Articulo", "Descripcion", "Fecha"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="121212", end_color="121212", fill_type="solid")

    for row in lecturas:
        fecha_valor = row["fecha_lectura"]
        if isinstance(fecha_valor, datetime):
            fecha_formateada = fecha_valor.strftime("%d/%m/%Y %H:%M")
        else:
            fecha_formateada = str(fecha_valor) if fecha_valor else ""

        ws.append([row["ean"], row["codigo_articulo"], row["descripcion"] or "", fecha_formateada])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"lecturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# === ARTICULOS ===

@app.route("/api/articulos/importar", methods=["POST"])
def importar_articulos():
    if not session.get("es_admin"):
        return jsonify({"success": False, "message": "No autorizado"}), 403

    if "archivo" not in request.files:
        return jsonify({"success": False, "message": "No se recibio ningun archivo"}), 400

    archivo = request.files["archivo"]
    if archivo.filename == "":
        return jsonify({"success": False, "message": "Nombre de archivo vacio"}), 400

    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "message": "El archivo debe ser Excel (.xlsx o .xls)"}), 400

    # Ajustable: 1000–5000 suele ir bien
    BATCH_SIZE = int(os.environ.get("IMPORT_BATCH_SIZE", "2000"))

    def norm_str(x):
        return str(x).strip() if x is not None else ""

    def norm_ean(x):
        # Soporta EAN como int/float/científico/string
        if x is None:
            return ""
        if isinstance(x, int):
            s = str(x)
        elif isinstance(x, float):
            s = str(int(x)) if x.is_integer() else str(x)
        else:
            s = str(x).strip()

        s = s.replace(" ", "").replace("\t", "").replace("\n", "")
        if "e" in s.lower():
            try:
                s = str(int(float(s)))
            except Exception:
                pass

        return "".join(ch for ch in s if ch.isdigit())

    try:
        from openpyxl import load_workbook

        # read_only=True reduce memoria y suele ir mejor con archivos grandes
        wb = load_workbook(archivo, data_only=True, read_only=True)
        ws = wb.active

        total_filas = 0
        descartadas = 0
        importados = 0
        batch = []

        insert_sql = """
            INSERT INTO articulos (codigo_articulo, descripcion, ean)
            VALUES (%s, %s, %s)
            ON CONFLICT (ean) DO NOTHING
        """

        with get_db() as conn:
            with conn.cursor() as cursor:
                cursor.execute("DELETE FROM articulos")

                for row in ws.iter_rows(min_row=2, values_only=True):
                    total_filas += 1

                    codigo_articulo = norm_str(row[0] if len(row) > 0 else None)
                    descripcion = norm_str(row[1] if len(row) > 1 else None)
                    ean = norm_ean(row[2] if len(row) > 2 else None)

                    if not codigo_articulo or not ean:
                        descartadas += 1
                        continue

                    batch.append((codigo_articulo, descripcion, ean))

                    if len(batch) >= BATCH_SIZE:
                        cursor.executemany(insert_sql, batch)
                        importados += len(batch)
                        batch.clear()

                # flush final
                if batch:
                    cursor.executemany(insert_sql, batch)
                    importados += len(batch)
                    batch.clear()

            conn.commit()

        return jsonify({
            "success": True,
            "message": f"Importacion OK. Filas leidas: {total_filas}. Importadas (intentadas): {importados}. Descartadas: {descartadas}. Batch: {BATCH_SIZE}."
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500


@app.route("/api/articulos/count", methods=["GET"])
def contar_articulos():
    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) AS total FROM articulos")
            count = cursor.fetchone()["total"]
    return jsonify({"count": count})


@app.route("/api/articulos/limpiar", methods=["DELETE"])
def limpiar_articulos():
    if not session.get("es_admin"):
        return jsonify({"success": False, "message": "No autorizado"}), 403

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM articulos")
        conn.commit()

    return jsonify({"success": True, "message": "Tabla maestra limpiada"})


# === USUARIOS (ADMIN) ===

@app.route("/api/admin/usuarios", methods=["GET"])
def obtener_usuarios():
    if not session.get("es_admin"):
        return jsonify({"success": False, "message": "No autorizado"}), 403

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT id, nombre_usuario, es_admin, fecha_creacion FROM usuarios ORDER BY fecha_creacion DESC"
            )
            usuarios = cursor.fetchall()

    return jsonify(usuarios)


@app.route("/api/admin/usuarios", methods=["POST"])
def crear_usuario():
    if not session.get("es_admin"):
        return jsonify({"success": False, "message": "No autorizado"}), 403

    data = request.json or {}
    nombre = (data.get("nombre_usuario") or "").strip()
    if not nombre:
        return jsonify({"success": False, "message": "Nombre de usuario requerido"}), 400

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO usuarios (nombre_usuario, es_admin)
                VALUES (%s, %s)
                ON CONFLICT (nombre_usuario) DO NOTHING
                RETURNING id
                """,
                (nombre, False),
            )
            nuevo = cursor.fetchone()
        conn.commit()

    if nuevo:
        return jsonify({"success": True, "message": "Usuario creado", "id": nuevo["id"]})

    return jsonify({"success": False, "message": "El usuario ya existe"}), 400


@app.route("/api/admin/usuarios/<int:usuario_id>", methods=["DELETE"])
def eliminar_usuario(usuario_id: int):
    if not session.get("es_admin"):
        return jsonify({"success": False, "message": "No autorizado"}), 403

    with get_db() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT nombre_usuario FROM usuarios WHERE id = %s", (usuario_id,))
            user = cursor.fetchone()

            if not user:
                return jsonify({"success": False, "message": "Usuario no encontrado"}), 404

            if user["nombre_usuario"] in ["admin", "henkobit"]:
                return jsonify({"success": False, "message": "No se puede eliminar este administrador"}), 400

            cursor.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
        conn.commit()

    return jsonify({"success": True, "message": "Usuario eliminado"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
